#!/usr/bin/env python3
"""
Categorize resources based on their description and services.
This is a fast, standalone script using keyword-based categorization.
"""

import argparse
import csv
from pathlib import Path
from typing import Dict, List
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill


# Valid resource categories
RESOURCE_CATEGORIES = [
    "Assessment & Diagnosis",
    "Crisis & Emergency",
    "Community Support",
    "Education Support",
    "Employment Support",
    "Housing Support",
    "Benefits Support",
    "Transport Support",
    "Autism Friendly Entertainment",
    "Local Support Services and Groups",
    "Employment and Skills Services",
    "National Autistic Society Branches",
    "Autism Friendly Sports Activities",
    "Special Needs Play Centres",
    "Special Interests and Hobbies",
    "Unknown Category",
]


def categorize_resource(row: Dict[str, str]) -> str:
    """
    Categorize a resource based on its name, description, and services.
    Uses keyword matching for fast, accurate categorization.
    """
    name = row.get("gmaps_name", "")
    description = row.get("description_short", "")
    services = row.get("specific_services", "")
    org_type = row.get("organization_type", "")
    
    # If no description and no services, return Unknown
    if (not description or description.strip().lower() in ["", "nan", "none", "null", "n/a", "not available"]) and \
       (not services or services.strip().lower() in ["", "nan", "none", "null", "n/a", "not available"]):
        return "Unknown Category"
    
    # Combine all text fields for analysis (giving more weight to services)
    name_lower = name.lower()
    desc_lower = description.lower()
    services_lower = (services or "").lower()
    
    # Weight services more heavily by repeating it
    combined = f"{name_lower} {desc_lower} {services_lower} {services_lower}"
    
    # National Autistic Society branches (check first for specificity)
    if "national autistic society" in combined:
        if "branch" in combined or "nas " in name_lower:
            return "National Autistic Society Branches"
    
    # Assessment & Diagnosis
    if any(keyword in combined for keyword in [
        "assessment", "diagnosis", "diagnostic", "evaluation", 
        "camhs", "screening", "testing", "psychiatric assessment"
    ]):
        return "Assessment & Diagnosis"
    
    # Crisis & Emergency  
    if any(keyword in combined for keyword in [
        "crisis", "emergency", "helpline", "24/7", "urgent",
        "immediate support", "samaritans", "crisis line"
    ]):
        return "Crisis & Emergency"
    
    # Employment (check skills-based first)
    employment_keywords = ["employment", "job", "work", "career", "employability", "vocational"]
    skills_keywords = ["skills training", "skills development", "apprenticeship", "work experience"]
    
    if any(keyword in combined for keyword in skills_keywords):
        return "Employment and Skills Services"
    elif any(keyword in combined for keyword in employment_keywords):
        return "Employment Support"
    
    # Sports (very specific keywords)
    if any(keyword in combined for keyword in [
        "sport", "swimming", "football", "cycling", "gymnastics", 
        "fitness", "physical activity", "swim", "athletics",
        "tennis", "basketball", "exercise", "trampolining"
    ]):
        return "Autism Friendly Sports Activities"
    
    # Entertainment & Arts
    if any(keyword in combined for keyword in [
        "theatre", "cinema", "music", "art", "drama", "dance",
        "entertainment", "museum", "gallery", "performance",
        "concert", "show", "festival"
    ]):
        return "Autism Friendly Entertainment"
    
    # Special interests & hobbies
    if any(keyword in combined for keyword in [
        "hobby", "hobbies", "club", "railway", "model", "crafts",
        "interests", "chess", "gaming", "computers", "lego"
    ]):
        return "Special Interests and Hobbies"
    
    # Play centres
    if any(keyword in combined for keyword in [
        "play centre", "playscheme", "playground", "play group", 
        "playgroup", "soft play", "adventure playground"
    ]):
        return "Special Needs Play Centres"
    
    # Education
    if any(keyword in combined for keyword in [
        "school", "education", "learning", "teaching", "tutoring",
        "sen support", "educational", "tutor", "teacher",
        "classroom", "lesson"
    ]):
        return "Education Support"
    
    # Housing
    if any(keyword in combined for keyword in [
        "housing", "accommodation", "residential", "living",
        "home", "residence", "supported living"
    ]):
        return "Housing Support"
    
    # Benefits
    if any(keyword in combined for keyword in [
        "benefit", "welfare", "dla", "pip", "financial support",
        "financial advice", "benefits advice", "universal credit"
    ]):
        return "Benefits Support"
    
    # Transport
    if any(keyword in combined for keyword in [
        "transport", "travel", "mobility", "accessible transport",
        "travel training", "tfl", "transport for london"
    ]):
        return "Transport Support"
    
    # Local support groups (parent groups, support groups, etc.)
    if any(keyword in combined for keyword in [
        "parent", "carer", "support group", "meetup", 
        "coffee morning", "forum", "parents", "carers",
        "peer support", "family support"
    ]):
        return "Local Support Services and Groups"
    
    # Community support (general - check last as it's broad)
    if any(keyword in combined for keyword in [
        "community", "support", "advice", "guidance", 
        "help", "service", "charity"
    ]):
        return "Community Support"
    
    # Default
    return "Unknown Category"


def read_input_file(file_path: str) -> List[Dict[str, str]]:
    """Read Excel or CSV file."""
    path = Path(file_path)
    
    if path.suffix.lower() in ['.xlsx', '.xls']:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                value = ws.cell(row_idx, col_idx).value
                row_data[header] = str(value) if value is not None else ""
            rows.append(row_data)
        return rows
    else:
        # CSV
        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except UnicodeDecodeError:
                continue
        raise Exception("Could not read file with any encoding")


def export_to_xlsx(data: List[Dict[str, str]], output_path: str, changed_indices: set):
    """Export data to Excel with highlighting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorized Resources"
    
    if not data:
        wb.save(output_path)
        return
    
    # Collect all unique column names
    all_headers = []
    for row in data:
        for key in row.keys():
            if key not in all_headers:
                all_headers.append(key)
    
    # Write headers
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Write data
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(all_headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Highlight categorized rows
            if row_idx - 2 in changed_indices and header == "resource_category":
                cell.fill = yellow_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Categorize neurodivergent support resources"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Input Excel or CSV file"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file (default: adds _categorized suffix)"
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing categories (default: only fill empty)"
    )
    
    args = parser.parse_args()
    
    # Read input
    print(f"Reading {args.input}...")
    rows = read_input_file(args.input)
    print(f"Loaded {len(rows)} rows")
    
    # Ensure resource_category column exists
    for row in rows:
        if "resource_category" not in row:
            row["resource_category"] = ""
    
    # Categorize
    print("\n📂 Categorizing resources...")
    categorized_count = 0
    changed_indices = set()
    
    for idx, row in enumerate(rows):
        existing_category = row.get("resource_category", "").strip()
        
        # Skip if already categorized and not overwriting
        if existing_category and not args.overwrite:
            continue
        
        category = categorize_resource(row)
        row["resource_category"] = category
        categorized_count += 1
        changed_indices.add(idx)
        
        if categorized_count % 100 == 0:
            print(f"  Categorized {categorized_count} resources...")
    
    print(f"\n✅ Categorized {categorized_count} resources")
    
    # Show category distribution
    categories = [row.get("resource_category", "Unknown Category") for row in rows]
    category_counts = Counter(categories)
    
    print("\n📊 Category Distribution:")
    print("="*70)
    for cat in RESOURCE_CATEGORIES:
        count = category_counts.get(cat, 0)
        percentage = (count / len(rows) * 100) if len(rows) > 0 else 0
        print(f"  {cat:45} {count:4} ({percentage:5.1f}%)")
    print("="*70)
    print(f"  {'TOTAL':45} {len(rows):4} (100.0%)")
    
    # Export
    if args.output:
        output_path = args.output
    else:
        input_path = Path(args.input)
        output_path = input_path.parent / f"{input_path.stem}_categorized.xlsx"
    
    print(f"\nExporting to {output_path}...")
    export_to_xlsx(rows, str(output_path), changed_indices)
    
    print(f"\n✅ Done! Output saved to: {output_path}")
    print("Categories are highlighted in yellow")


if __name__ == "__main__":
    main()

