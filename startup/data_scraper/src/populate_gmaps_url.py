#!/usr/bin/env python3
"""
Populate missing gmaps_url by searching Google Places API.

This script:
1. Reads an Excel/CSV file
2. Finds rows with missing gmaps_url
3. Searches Google Places API using name + address (or coordinates)
4. Gets place_id and constructs proper gmaps_url
5. Updates the file with new URLs
"""

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, Optional, Tuple

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# Try to load .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


def get_google_api_key() -> str:
    """Get Google API key from environment."""
    key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not key:
        raise RuntimeError(
            "Google API key not found. Set GOOGLE_API_KEY or GEMINI_API_KEY in .env file"
        )
    return key


def search_place_by_text(query: str, api_key: str) -> Optional[Dict]:
    """
    Search for a place using Google Places API Text Search.
    
    Returns place details including place_id.
    """
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {
        "query": query,
        "key": api_key,
    }
    
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if data.get("status") == "OK" and data.get("results"):
            # Return first result
            return data["results"][0]
        else:
            print(f"  ⚠️  No results: {data.get('status')}", file=sys.stderr)
            return None
            
    except Exception as e:
        print(f"  ❌ API error: {e}", file=sys.stderr)
        return None


def search_place_by_coordinates(lat: float, lng: float, name: str, api_key: str) -> Optional[Dict]:
    """
    Search for a place using coordinates with Google Places API Nearby Search.
    """
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": 50,  # 50 meters
        "name": name,
        "key": api_key,
    }
    
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if data.get("status") == "OK" and data.get("results"):
            return data["results"][0]
        else:
            # Fallback to text search if nearby fails
            return search_place_by_text(f"{name} {lat},{lng}", api_key)
            
    except Exception as e:
        print(f"  ❌ API error: {e}", file=sys.stderr)
        return None


def construct_maps_url(place_id: str, use_directions: bool = False) -> str:
    """
    Construct Google Maps URL from place_id.
    
    Two formats:
    1. Direct link: https://maps.google.com/?cid={cid}
    2. Directions link: https://www.google.com/maps/dir/?api=1&destination_place_id={place_id}
    """
    if use_directions:
        return f"https://www.google.com/maps/dir/?api=1&destination_place_id={place_id}"
    else:
        return f"https://www.google.com/maps/place/?q=place_id:{place_id}"


def construct_coordinate_url(lat: float, lng: float) -> str:
    """Construct a simple coordinate-based Maps URL."""
    return f"https://www.google.com/maps?q={lat},{lng}"


def is_field_empty(value) -> bool:
    """Check if a field is empty or contains placeholder text."""
    if value is None:
        return True
    if isinstance(value, str):
        value = value.strip().lower()
        return not value or value in ["", "nan", "none", "null", "n/a", "not available", "unknown"]
    return False


def is_valid_place_id(place_id: str) -> bool:
    """
    Check if a place_id looks like a valid Google Place ID.
    Real Place IDs are alphanumeric, usually start with 'ChIJ', and are 27+ chars.
    """
    if not place_id or not isinstance(place_id, str):
        return False
    
    place_id = place_id.strip()
    
    # Filter out obvious placeholders
    if place_id.lower() in ["", "nan", "none", "null", "n/a", "unknown", "not available"]:
        return False
    
    # Reject pure numbers (placeholders like "1", "2", "3")
    if place_id.isdigit():
        return False
    
    # Real Place IDs are typically 27+ characters long
    if len(place_id) < 15:
        return False
    
    # Should be alphanumeric (with underscores/hyphens allowed)
    if not all(c.isalnum() or c in '_-' for c in place_id):
        return False
    
    return True


def populate_gmaps_url(
    row: Dict[str, str],
    api_key: str,
    use_directions: bool = False,
    use_api: bool = True,
    rate_limit_delay: float = 0.1
) -> Tuple[Optional[str], Optional[str], str]:
    """
    Populate gmaps_url for a single row.
    
    Returns: (gmaps_url, place_id_if_new, status_message)
    """
    # Check if already has URL
    if not is_field_empty(row.get("gmaps_url")):
        return None, None, "already has URL"
    
    # Strategy 1: If we already have VALID place_id, just construct URL
    existing_place_id = row.get("gmaps_place_id", "").strip()
    if is_valid_place_id(existing_place_id):
        url = construct_maps_url(existing_place_id, use_directions)
        return url, None, "constructed from existing place_id"
    
    # Strategy 2: If we have coordinates, try coordinate-based search or simple URL
    lat = row.get("gmaps_latitude", "").strip()
    lng = row.get("gmaps_longitude", "").strip()
    name = row.get("gmaps_name", "").strip()
    
    if lat and lng and not is_field_empty(lat) and not is_field_empty(lng):
        try:
            lat_f = float(lat)
            lng_f = float(lng)
            
            if use_api:
                # Try to find place via API
                time.sleep(rate_limit_delay)
                place_data = search_place_by_coordinates(lat_f, lng_f, name, api_key)
                if place_data:
                    place_id = place_data.get("place_id")
                    url = construct_maps_url(place_id, use_directions)
                    return url, place_id, "found via coordinates search"
            
            # Fallback: Use simple coordinate URL
            url = construct_coordinate_url(lat_f, lng_f)
            return url, None, "coordinate-based URL"
            
        except ValueError:
            pass
    
    # Strategy 3: Search by name + address
    if name and not is_field_empty(name):
        address = row.get("gmaps_formatted_address", "").strip()
        query = f"{name} {address}" if address and not is_field_empty(address) else name
        
        if use_api:
            time.sleep(rate_limit_delay)
            place_data = search_place_by_text(query, api_key)
            if place_data:
                place_id = place_data.get("place_id")
                url = construct_maps_url(place_id, use_directions)
                return url, place_id, "found via text search"
    
    return None, None, "could not find place"


def read_input_file(file_path: str) -> Tuple[list, list]:
    """Read Excel or CSV file. Returns (headers, rows)."""
    path = Path(file_path)
    
    if path.suffix.lower() in ['.xlsx', '.xls']:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                row_data[header] = ws.cell(row_idx, col_idx).value or ""
            rows.append(row_data)
        return headers, rows
    else:
        # CSV
        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    headers = list(rows[0].keys()) if rows else []
                    return headers, rows
            except UnicodeDecodeError:
                continue
        raise Exception("Could not read file with any encoding")


def export_to_xlsx(
    data: list,
    output_path: str,
    headers: list,
    changed_rows: set
) -> None:
    """Export data to Excel with highlighting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Updated URLs"
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Write data
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Highlight changed rows
            if row_idx - 2 in changed_rows:
                if header in ['gmaps_url', 'gmaps_place_id']:
                    cell.fill = yellow_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Populate missing gmaps_url using Google Places API"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Input Excel or CSV file"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output file (default: adds _with_urls suffix)"
    )
    parser.add_argument(
        "--use-directions",
        action="store_true",
        help="Use directions URL format instead of place URL format"
    )
    parser.add_argument(
        "--no-api",
        action="store_true",
        help="Don't call API, only construct URLs from existing data"
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Maximum rows to process (for testing)"
    )
    parser.add_argument(
        "--rate-limit",
        type=float,
        default=0.1,
        help="Delay between API calls in seconds (default: 0.1)"
    )
    
    args = parser.parse_args()
    
    # Get API key
    try:
        api_key = get_google_api_key()
    except RuntimeError as e:
        if not args.no_api:
            print(f"❌ {e}", file=sys.stderr)
            sys.exit(1)
        api_key = None
    
    # Read input
    print(f"Reading {args.input}...")
    headers, rows = read_input_file(args.input)
    print(f"Loaded {len(rows)} rows")
    
    if args.max_rows:
        rows = rows[:args.max_rows]
        print(f"Limited to {len(rows)} rows for testing")
    
    # Check which rows need URLs
    rows_needing_urls = []
    for idx, row in enumerate(rows):
        if is_field_empty(row.get("gmaps_url")):
            rows_needing_urls.append(idx)
    
    print(f"\nFound {len(rows_needing_urls)} rows without gmaps_url")
    
    if not rows_needing_urls:
        print("✅ All rows already have gmaps_url!")
        return
    
    # Ensure gmaps_url and gmaps_place_id columns exist
    if "gmaps_url" not in headers:
        headers.append("gmaps_url")
    if "gmaps_place_id" not in headers:
        headers.append("gmaps_place_id")
    
    # Process rows
    updated_count = 0
    failed_count = 0
    changed_rows = set()
    
    print(f"\nProcessing {len(rows_needing_urls)} rows...\n")
    
    for count, idx in enumerate(rows_needing_urls, 1):
        row = rows[idx]
        name = row.get("gmaps_name", "Unknown")
        
        print(f"[{count}/{len(rows_needing_urls)}] {name[:50]}...", end=" ")
        
        url, new_place_id, status = populate_gmaps_url(
            row,
            api_key,
            use_directions=args.use_directions,
            use_api=not args.no_api,
            rate_limit_delay=args.rate_limit
        )
        
        if url:
            row["gmaps_url"] = url
            if new_place_id and is_field_empty(row.get("gmaps_place_id")):
                row["gmaps_place_id"] = new_place_id
            changed_rows.add(idx)
            updated_count += 1
            print(f"✅ {status}")
        else:
            failed_count += 1
            print(f"❌ {status}")
    
    # Export results
    if args.output:
        output_path = args.output
    else:
        input_path = Path(args.input)
        output_path = input_path.parent / f"{input_path.stem}_with_urls{input_path.suffix}"
    
    print(f"\nExporting to {output_path}...")
    
    if str(output_path).endswith('.xlsx'):
        export_to_xlsx(rows, str(output_path), headers, changed_rows)
    else:
        # CSV export
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
    
    # Summary
    print("\n" + "="*70)
    print("SUMMARY")
    print("="*70)
    print(f"Total rows processed:    {len(rows_needing_urls)}")
    print(f"✅ Successfully updated: {updated_count}")
    print(f"❌ Failed:               {failed_count}")
    print(f"\nOutput saved to: {output_path}")
    print("="*70)


if __name__ == "__main__":
    main()

