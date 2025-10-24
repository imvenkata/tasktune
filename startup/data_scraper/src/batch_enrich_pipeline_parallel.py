#!/usr/bin/env python3
"""
Parallel batch pipeline to enrich service center information
Much faster than the sequential version using concurrent processing.
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from threading import Lock, Semaphore
from typing import Dict, List, Optional, Tuple

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    env_path = Path(__file__).parent / '.env'
    load_dotenv(dotenv_path=env_path)
except ImportError:
    pass  # python-dotenv not installed, will use system env vars

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None  # Will be checked if --populate-urls is used


# Valid resource categories
RESOURCE_CATEGORIES = [
    "Assessment & Diagnosis",
    "Crisis & Emergency",
    "Community Support",
    "Education Support",
    "Employment Support",
    "Housing Support",
    "Benefits Support",
    "Transport Support",
    "Autism Friendly Entertainment",
    "Local Support Services and Groups",
    "Employment and Skills Services",
    "National Autistic Society Branches",
    "Autism Friendly Sports Activities",
    "Special Needs Play Centres",
    "Special Interests and Hobbies",
    "Unknown Category",
]

# Column mapping from web_llm_extract output to enriched_resources.csv columns
COLUMN_MAPPING = {
    "center_name": "gmaps_name",
    "website_url": "gmaps_website",
    "description_short": "description_short",
    "age_range": "age_range",
    "conditions_supported": "conditions_supported",
    "specific_services": "specific_services",
    "organization_type": "organization_type",
    "contact_info.phone": "gmaps_phone",
    "contact_info.email": None,
    "contact_info.address": "gmaps_formatted_address",
    "address_components.postal_code": "gmaps_addr_postal_code",
    "address_components.postal_town": "gmaps_addr_postal_town",
    "address_components.admin_area_level_1": "gmaps_addr_admin_area_level_1",
    "address_components.admin_area_level_2": "gmaps_addr_admin_area_level_2",
    # Neurodivergent relevance validation
    "is_neurodivergent_related": "is_neurodivergent_related",
    "neurodivergent_relevance_score": "neurodivergent_relevance_score",
    "neurodivergent_focus": "neurodivergent_focus",
    # Resource categorization
    "resource_category": "resource_category",
}

# Thread-safe locks
cache_lock = Lock()
stats_lock = Lock()

# Rate limiting (15 requests per minute for Gemini free tier)
class RateLimiter:
    def __init__(self, max_per_minute=15):
        self.max_per_minute = max_per_minute
        self.requests = []
        self.lock = Lock()
    
    def wait_if_needed(self):
        """Wait if we've hit the rate limit."""
        with self.lock:
            now = time.time()
            # Remove requests older than 1 minute
            self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            if len(self.requests) >= self.max_per_minute:
                # Wait until the oldest request is 60 seconds old
                sleep_time = 60 - (now - self.requests[0]) + 0.1
                if sleep_time > 0:
                    time.sleep(sleep_time)
                    # Clean up again after sleeping
                    now = time.time()
                    self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            self.requests.append(now)

rate_limiter = RateLimiter(max_per_minute=15)


def is_field_empty(value: any) -> bool:
    """Check if a field is empty or missing."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() in ("", "Not specified", "N/A", "Unknown")
    if isinstance(value, list):
        return len(value) == 0
    return False


def needs_enrichment(row: Dict[str, str], fields_to_check: List[str]) -> bool:
    """Check if a row needs enrichment based on missing fields."""
    for field in fields_to_check:
        if is_field_empty(row.get(field, "")):
            return True
    return False


def extract_nested_value(data: Dict, key_path: str) -> any:
    """Extract value from nested dictionary using dot notation."""
    keys = key_path.split(".")
    value = data
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return None
    return value


def call_web_llm_extract(
    center_name: str,
    website_url: str,
    backend: str,
    model: Optional[str],
    enhance_with_websearch: bool,
    cache_dir: str,
    use_rate_limit: bool = True,
    **kwargs
) -> Optional[Dict]:
    """Call web_llm_extract.py to extract information."""
    # Check cache first (thread-safe)
    cache_file = Path(cache_dir) / f"{center_name.replace('/', '_').replace(' ', '_')[:100]}.json"
    if cache_file.exists():
        with cache_lock:
            try:
                with open(cache_file, "r", encoding="utf-8") as f:
                    cached_data = json.load(f)
                    
                    # Add neurodivergent validation fields if missing (for old cache)
                    if "is_neurodivergent_related" not in cached_data:
                        cached_data["is_neurodivergent_related"] = True  # Assume neurodivergent if in this dataset
                        cached_data["neurodivergent_relevance_score"] = "Not validated (old cache)"
                        cached_data["neurodivergent_focus"] = "Validation not available (cached before validation feature)"
                    
                    return {"cached": True, "data": cached_data}
            except (OSError, json.JSONDecodeError):
                pass
    
    # Apply rate limiting for new API calls (skip for cached)
    if use_rate_limit and backend in ("gemini", "openai"):
        rate_limiter.wait_if_needed()
    
    # Build command
    cmd = [
        sys.executable,
        "web_llm_extract.py",
        "--center-name", center_name,
        "--url", website_url,
        "--backend", backend,
    ]
    
    if model:
        cmd.extend(["--model", model])
    
    if enhance_with_websearch:
        cmd.append("--enhance-with-websearch")
    
    # Add backend-specific args
    if backend == "ollama" and kwargs.get("ollama_url"):
        cmd.extend(["--ollama-url", kwargs["ollama_url"]])
    if backend == "ollama" and kwargs.get("ollama_auth"):
        cmd.extend(["--ollama-auth", kwargs["ollama_auth"]])
    if backend == "openai" and kwargs.get("openai_key"):
        cmd.extend(["--openai-key", kwargs["openai_key"]])
    if backend == "gemini" and kwargs.get("gemini_key"):
        cmd.extend(["--gemini-key", kwargs["gemini_key"]])
    
    try:
        # Pass current environment to subprocess (includes .env vars)
        env = os.environ.copy()
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
            env=env,
            cwd=str(Path(__file__).parent)
        )
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            
            # Cache the result (thread-safe)
            with cache_lock:
                cache_file.parent.mkdir(parents=True, exist_ok=True)
                with open(cache_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            
            return {"cached": False, "data": data}
        else:
            return None
    except subprocess.TimeoutExpired:
        return None
    except (json.JSONDecodeError, OSError):
        return None


def merge_data(original_row: Dict[str, str], extracted_data: Dict) -> Tuple[Dict[str, str], List[str]]:
    """Merge extracted data into original row, only filling missing fields."""
    updated_row = original_row.copy()
    changes = []
    
    for extract_key, csv_column in COLUMN_MAPPING.items():
        if csv_column is None:
            continue
        
        extracted_value = extract_nested_value(extracted_data, extract_key)
        if extracted_value is None or is_field_empty(extracted_value):
            continue
        
        # Get original value, initialize new column if doesn't exist
        original_value = updated_row.get(csv_column, "")
        
        # Add new column or fill empty existing column
        if is_field_empty(original_value) or csv_column not in updated_row:
            if isinstance(extracted_value, list):
                extracted_value = ", ".join(str(v) for v in extracted_value if v)
            elif isinstance(extracted_value, dict):
                extracted_value = json.dumps(extracted_value)
            
            updated_row[csv_column] = str(extracted_value)
            changes.append(csv_column)
    
    return updated_row, changes


def process_single_center(
    task: Tuple[int, int, Dict[str, str]],
    args,
    total_tasks: int
) -> Tuple[int, Optional[Dict[str, str]], List[str], str]:
    """Process a single center (designed for parallel execution)."""
    process_idx, original_idx, row = task
    
    center_name = row.get("gmaps_name", "Unknown")
    website = row.get("gmaps_website", "").strip()
    
    status_msg = f"[{process_idx}/{total_tasks}] {center_name[:50]}"
    
    if not website:
        return original_idx, None, [], f"{status_msg}: No website"
    
    try:
        result = call_web_llm_extract(
            center_name=center_name,
            website_url=website,
            backend=args.backend,
            model=args.model,
            enhance_with_websearch=args.enhance_with_websearch,
            cache_dir=args.cache_dir,
            ollama_url=args.ollama_url,
            ollama_auth=args.ollama_auth,
            openai_key=args.openai_key,
            gemini_key=args.gemini_key
        )
        
        if result and "data" in result and "error" not in result["data"]:
            updated_row, changes = merge_data(row, result["data"])
            cached = "✓ cached" if result.get("cached") else "✓ new"
            return original_idx, updated_row, changes, f"{status_msg}: {cached} - {len(changes)} fields"
        else:
            return original_idx, None, [], f"{status_msg}: Failed"
    
    except Exception as e:
        return original_idx, None, [], f"{status_msg}: Error - {str(e)[:50]}"


def clean_text_for_excel(text: str) -> str:
    """Clean text to handle encoding issues for Excel."""
    if not isinstance(text, str):
        return text
    
    # First, handle common patterns from corrupted data
    # Pattern: "9:00?AM?�?5:00?PM" -> "9:00 AM - 5:00 PM"
    import re
    text = re.sub(r'(\d{1,2}:\d{2})\?([AP]M)\?[�\ufffd]?\?(\d{1,2}:\d{2})\?([AP]M)', 
                  r'\1 \2 - \3 \4', text)
    
    # Replace common problematic characters
    replacements = {
        '\u2013': '-',  # en dash
        '\u2014': '-',  # em dash
        '\u2018': "'",  # left single quote
        '\u2019': "'",  # right single quote
        '\u201c': '"',  # left double quote
        '\u201d': '"',  # right double quote
        '\u2026': '...',  # ellipsis
        '\xa0': ' ',  # non-breaking space
        '\u00a0': ' ',  # non-breaking space (alternative)
        '\ufffd': '',  # replacement character (�)
        '�': '',  # replacement character (alternative)
        '\x00': '',  # null byte
    }
    
    for old, new in replacements.items():
        text = text.replace(old, new)
    
    # Remove standalone ? that aren't part of real text (but preserve in URLs)
    # Only remove ? if it's followed by letters/spaces (garbage pattern like "9:00?AM")
    # Don't remove ? that's part of URL query strings
    if 'http' not in text.lower():
        # Not a URL, safe to remove all ?
        text = re.sub(r'\?+', ' ', text)
    else:
        # URL detected, only remove ? in garbage patterns, not in query strings
        # Remove patterns like "?AM" or "?PM" but keep "?q=" or "?key="
        text = re.sub(r'\?([AP]M)', r' \1', text)  # Fix time garbage
        text = re.sub(r'(\d)\?([A-Z][a-z])', r'\1 \2', text)  # Fix other garbage
    
    # Clean up multiple spaces
    text = re.sub(r' +', ' ', text)
    
    # Remove any remaining non-printable characters except newlines/tabs
    text = ''.join(char if char.isprintable() or char in '\n\r\t' else ' ' for char in text)
    
    return text.strip()


def is_valid_place_id(place_id: str) -> bool:
    """
    Check if a place_id looks like a valid Google Place ID.
    Real Place IDs are alphanumeric, usually start with 'ChIJ', and are 27+ chars.
    """
    if not place_id or not isinstance(place_id, str):
        return False
    
    place_id = place_id.strip()
    
    # Filter out obvious placeholders
    if place_id.lower() in ["", "nan", "none", "null", "n/a", "unknown", "not available"]:
        return False
    
    # Reject pure numbers (placeholders like "1", "2", "3")
    if place_id.isdigit():
        return False
    
    # Real Place IDs are typically 27+ characters long
    if len(place_id) < 15:
        return False
    
    # Should be alphanumeric (with underscores/hyphens allowed)
    if not all(c.isalnum() or c in '_-' for c in place_id):
        return False
    
    return True


def categorize_resource(row: Dict[str, str], backend: str = "gemini") -> str:
    """
    Categorize a resource based on its description and other fields.
    Returns a category from RESOURCE_CATEGORIES.
    """
    name = row.get("gmaps_name", "")
    description = row.get("description_short", "")
    services = row.get("specific_services", "")
    org_type = row.get("organization_type", "")
    
    # If no description, return Unknown
    if not description or description.strip().lower() in ["", "nan", "none", "null", "n/a", "not available"]:
        return "Unknown Category"
    
    # Use simple keyword-based categorization for speed
    name_lower = name.lower()
    desc_lower = description.lower()
    services_lower = (services or "").lower()
    combined = f"{name_lower} {desc_lower} {services_lower}"
    
    # National Autistic Society branches
    if "national autistic society" in combined and ("branch" in combined or "nas " in combined):
        return "National Autistic Society Branches"
    
    # Assessment & Diagnosis
    if any(keyword in combined for keyword in ["assessment", "diagnosis", "diagnostic", "evaluation", "camhs"]):
        return "Assessment & Diagnosis"
    
    # Crisis & Emergency  
    if any(keyword in combined for keyword in ["crisis", "emergency", "helpline", "24/7", "urgent"]):
        return "Crisis & Emergency"
    
    # Employment
    if any(keyword in combined for keyword in ["employment", "job", "work", "career", "skills training", "employability"]):
        if "skills" in combined:
            return "Employment and Skills Services"
        return "Employment Support"
    
    # Sports
    if any(keyword in combined for keyword in ["sport", "swimming", "football", "cycling", "gymnastics", "fitness", "physical activity"]):
        return "Autism Friendly Sports Activities"
    
    # Entertainment & Arts
    if any(keyword in combined for keyword in ["theatre", "cinema", "music", "art", "drama", "dance", "entertainment", "museum", "gallery"]):
        return "Autism Friendly Entertainment"
    
    # Special interests & hobbies
    if any(keyword in combined for keyword in ["hobby", "hobbies", "club", "railway", "model", "crafts"]):
        return "Special Interests and Hobbies"
    
    # Play centres
    if any(keyword in combined for keyword in ["play centre", "playscheme", "playground", "play group", "playgroup"]):
        return "Special Needs Play Centres"
    
    # Education
    if any(keyword in combined for keyword in ["school", "education", "learning", "teaching", "tutoring", "sen support"]):
        return "Education Support"
    
    # Housing
    if any(keyword in combined for keyword in ["housing", "accommodation", "residential", "living"]):
        return "Housing Support"
    
    # Benefits
    if any(keyword in combined for keyword in ["benefit", "welfare", "dla", "pip", "financial support"]):
        return "Benefits Support"
    
    # Transport
    if any(keyword in combined for keyword in ["transport", "travel", "mobility", "accessible transport"]):
        return "Transport Support"
    
    # Local support groups (parent groups, support groups, etc.)
    if any(keyword in combined for keyword in ["parent", "carer", "support group", "meetup", "coffee morning", "forum"]):
        return "Local Support Services and Groups"
    
    # Community support (general)
    if any(keyword in combined for keyword in ["community", "support", "advice", "guidance", "help"]):
        return "Community Support"
    
    # Default
    return "Unknown Category"


def populate_missing_urls(rows: List[Dict[str, str]], use_api: bool = True) -> Tuple[int, int]:
    """
    Populate missing gmaps_url by constructing from place_id or searching API.
    Returns (updated_count, failed_count).
    """
    if not requests and use_api:
        print("⚠️  requests library not installed. Install with: pip install requests", file=sys.stderr)
        use_api = False
    
    # Get API key if needed
    api_key = None
    if use_api:
        api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
        if not api_key:
            print("⚠️  Google API key not found. Only constructing URLs from existing place_id", file=sys.stderr)
            use_api = False
    
    updated_count = 0
    failed_count = 0
    
    for row in rows:
        # Skip if already has URL
        existing_url = row.get("gmaps_url", "").strip()
        if existing_url and existing_url.lower() not in ["", "nan", "none", "null", "n/a"]:
            continue
        
        # Strategy 1: Construct from existing VALID place_id
        place_id = row.get("gmaps_place_id", "").strip()
        if is_valid_place_id(place_id):
            row["gmaps_url"] = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
            updated_count += 1
            continue
        
        # Strategy 2: Use coordinates to create simple URL
        lat = row.get("gmaps_latitude", "").strip()
        lng = row.get("gmaps_longitude", "").strip()
        if lat and lng:
            try:
                lat_f = float(lat)
                lng_f = float(lng)
                row["gmaps_url"] = f"https://www.google.com/maps?q={lat_f},{lng_f}"
                updated_count += 1
                continue
            except ValueError:
                pass
        
        # Strategy 3: Could search API here (but skipping to avoid API costs)
        # The populate_gmaps_url.py script can be run separately for deeper search
        
        failed_count += 1
    
    return updated_count, failed_count


def export_to_xlsx(data: List[Dict[str, str]], output_path: str, highlight_changes: Dict[int, List[str]]) -> None:
    """Export data to Excel with highlighting for changed cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Resources"
    
    if not data:
        wb.save(output_path)
        return
    
    # Collect all unique column names from all rows (to include new columns)
    all_headers = []
    for row in data:
        for key in row.keys():
            if key not in all_headers:
                all_headers.append(key)
    
    headers = all_headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            # Clean text to handle encoding issues
            if isinstance(value, str):
                value = clean_text_for_excel(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx - 2 in highlight_changes and header in highlight_changes[row_idx - 2]:
                cell.fill = yellow_fill
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Parallel batch pipeline to enrich service center information (MUCH FASTER!)"
    )
    parser.add_argument("--input", default="data/input/to_be_normalised/enriched_resources.csv")
    parser.add_argument("--output", default=None)
    parser.add_argument("--backend", choices=["ollama", "openai", "gemini"], default="gemini")
    parser.add_argument("--model", default=None)
    parser.add_argument("--enhance-with-websearch", action="store_true")
    parser.add_argument("--max-rows", type=int, default=None)
    parser.add_argument("--start-row", type=int, default=0)
    parser.add_argument("--cache-dir", default=".cache/llm_extractions")
    parser.add_argument("--fields-to-check", default="description_short,age_range,organization_type")
    parser.add_argument("--skip-no-website", action="store_true")
    parser.add_argument("--workers", type=int, default=10, help="Number of parallel workers (default: 10)")
    parser.add_argument("--rate-limit", type=int, default=15, help="Max API requests per minute (default: 15 for Gemini free tier)")
    parser.add_argument("--populate-urls", action="store_true", help="Populate missing gmaps_url from place_id or coordinates before enrichment")
    parser.add_argument("--categorize", action="store_true", help="Automatically categorize resources after enrichment")
    parser.add_argument("--ollama-url", default=None)
    parser.add_argument("--ollama-auth", default=None)
    parser.add_argument("--openai-key", default=None)
    parser.add_argument("--gemini-key", default=None)
    
    args = parser.parse_args()
    
    # Update rate limiter with user setting
    global rate_limiter
    rate_limiter = RateLimiter(max_per_minute=args.rate_limit)
    
    print(f"🚀 Parallel Processing with {args.workers} workers")
    print(f"📊 Rate limit: {args.rate_limit} requests/minute")
    print(f"Reading {args.input}...")
    
    try:
        # Try UTF-8 first, then fall back to other encodings
        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        rows = None
        
        for encoding in encodings_to_try:
            try:
                with open(args.input, "r", encoding=encoding, errors='replace') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                break
            except UnicodeDecodeError:
                continue
        
        if rows is None:
            raise Exception("Could not read CSV with any encoding")
            
    except Exception as e:
        print(f"Error reading CSV: {e}", file=sys.stderr)
        sys.exit(1)
    
    print(f"Loaded {len(rows)} rows")
    
    # Populate missing URLs if requested
    if args.populate_urls:
        print("\n🔗 Populating missing gmaps_url...")
        updated, failed = populate_missing_urls(rows, use_api=False)
        if updated > 0:
            print(f"  ✅ Populated {updated} URLs")
        if failed > 0:
            print(f"  ⚠️  {failed} URLs could not be populated")
    
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"data/output/enriched_resources_parallel_{timestamp}.xlsx"
    
    fields_to_check = [f.strip() for f in args.fields_to_check.split(",")]
    tasks = []
    
    for idx, row in enumerate(rows):
        if args.start_row and idx < args.start_row:
            continue
        if args.max_rows and len(tasks) >= args.max_rows:
            break
        
        website = row.get("gmaps_website", "").strip()
        if args.skip_no_website and not website:
            continue
        
        if needs_enrichment(row, fields_to_check):
            tasks.append((len(tasks) + 1, idx, row))
    
    print(f"Found {len(tasks)} rows that need enrichment")
    
    updated_rows = rows.copy()
    highlight_changes = {}
    stats = {"processed": 0, "enriched": 0, "failed": 0, "cached": 0}
    
    if not tasks:
        if args.categorize:
            print("No rows need enrichment, but will categorize existing data...\n")
        else:
            print("No rows need enrichment. Exiting.")
            return
    else:
        print(f"Processing with {args.workers} parallel workers...\n")
    
    start_time = time.time()
    
    # Process in parallel (only if there are tasks)
    if tasks:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {
                executor.submit(process_single_center, task, args, len(tasks)): task
                for task in tasks
            }
            
            for future in as_completed(futures):
                original_idx, updated_row, changes, status_msg = future.result()
                
                print(status_msg)
                
                if updated_row:
                    updated_rows[original_idx] = updated_row
                    with stats_lock:
                        stats["processed"] += 1
                        if "cached" in status_msg:
                            stats["cached"] += 1
                        if changes:
                            highlight_changes[original_idx] = changes
                            stats["enriched"] += 1
                else:
                    with stats_lock:
                        if "No website" not in status_msg:
                            stats["failed"] += 1
    
    elapsed = time.time() - start_time
    
    # Categorize resources if requested
    if args.categorize:
        print("\n📂 Categorizing resources...")
        categorized_count = 0
        for row in updated_rows:
            # Only categorize if not already categorized
            if not row.get("resource_category") or row.get("resource_category", "").strip() == "":
                category = categorize_resource(row, backend=args.backend)
                row["resource_category"] = category
                categorized_count += 1
                if categorized_count % 50 == 0:
                    print(f"  Categorized {categorized_count} resources...")
        
        print(f"  ✅ Categorized {categorized_count} resources")
        
        # Show category distribution
        from collections import Counter
        categories = [row.get("resource_category", "Unknown Category") for row in updated_rows]
        category_counts = Counter(categories)
        print("\n  Category Distribution:")
        for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
            print(f"    {cat}: {count}")
    
    print(f"\nExporting to {args.output}...")
    export_to_xlsx(updated_rows, args.output, highlight_changes)
    
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total rows in input:     {len(rows)}")
    print(f"Rows needing enrichment: {len(tasks)}")
    print(f"Successfully processed:  {stats['processed']}")
    print(f"  - From cache:          {stats['cached']}")
    print(f"  - Newly enriched:      {stats['processed'] - stats['cached']}")
    print(f"Successfully enriched:   {stats['enriched']}")
    print(f"Failed:                  {stats['failed']}")
    print(f"\nTime elapsed:            {elapsed:.1f} seconds")
    print(f"Average per center:      {elapsed/len(tasks):.2f} seconds")
    print(f"\nOutput saved to: {args.output}")
    print("Changed cells are highlighted in yellow")
    print("="*60)


if __name__ == "__main__":
    main()

