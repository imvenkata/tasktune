#!/usr/bin/env python3
"""
Batch pipeline to enrich service center information using web_llm_extract.py
Reads from enriched_resources.csv and updates missing information.
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    env_path = Path(__file__).parent / '.env'
    load_dotenv(dotenv_path=env_path)
except ImportError:
    pass  # python-dotenv not installed, will use system env vars

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Column mapping from web_llm_extract output to enriched_resources.csv columns
COLUMN_MAPPING = {
    "center_name": "gmaps_name",
    "website_url": "gmaps_website",
    "description_short": "description_short",
    "age_range": "age_range",
    "conditions_supported": "conditions_supported",
    "specific_services": "specific_services",
    "organization_type": "organization_type",
    # Contact info mapping
    "contact_info.phone": "gmaps_phone",
    "contact_info.email": None,  # Not in original CSV
    "contact_info.address": "gmaps_formatted_address",
    # Address components mapping
    "address_components.postal_code": "gmaps_addr_postal_code",
    "address_components.postal_town": "gmaps_addr_postal_town",
    "address_components.admin_area_level_1": "gmaps_addr_admin_area_level_1",
    "address_components.admin_area_level_2": "gmaps_addr_admin_area_level_2",
    # Neurodivergent relevance validation
    "is_neurodivergent_related": "is_neurodivergent_related",
    "neurodivergent_relevance_score": "neurodivergent_relevance_score",
    "neurodivergent_focus": "neurodivergent_focus",
}


def is_field_empty(value: any) -> bool:
    """Check if a field is empty or missing."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() in ("", "Not specified", "N/A", "Unknown")
    if isinstance(value, list):
        return len(value) == 0
    return False


def needs_enrichment(row: Dict[str, str], fields_to_check: List[str]) -> bool:
    """Check if a row needs enrichment based on missing fields."""
    for field in fields_to_check:
        if is_field_empty(row.get(field, "")):
            return True
    return False


def extract_nested_value(data: Dict, key_path: str) -> any:
    """Extract value from nested dictionary using dot notation."""
    keys = key_path.split(".")
    value = data
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return None
    return value


def call_web_llm_extract(
    center_name: str,
    website_url: str,
    backend: str,
    model: Optional[str],
    enhance_with_websearch: bool,
    cache_dir: str,
    **kwargs
) -> Optional[Dict]:
    """Call web_llm_extract.py to extract information."""
    # Check cache first
    cache_file = Path(cache_dir) / f"{center_name.replace('/', '_').replace(' ', '_')[:100]}.json"
    if cache_file.exists():
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached_data = json.load(f)
                
                # Add neurodivergent validation fields if missing (for old cache)
                if "is_neurodivergent_related" not in cached_data:
                    cached_data["is_neurodivergent_related"] = True  # Assume neurodivergent if in this dataset
                    cached_data["neurodivergent_relevance_score"] = "Not validated (old cache)"
                    cached_data["neurodivergent_focus"] = "Validation not available (cached before validation feature)"
                
                print("  ✓ Using cached data", file=sys.stderr)
                return cached_data
        except (OSError, json.JSONDecodeError):
            pass
    
    # Build command
    cmd = [
        sys.executable,
        "web_llm_extract.py",
        "--center-name", center_name,
        "--url", website_url,
        "--backend", backend,
    ]
    
    if model:
        cmd.extend(["--model", model])
    
    if enhance_with_websearch:
        cmd.append("--enhance-with-websearch")
    
    # Add backend-specific args
    if backend == "ollama" and kwargs.get("ollama_url"):
        cmd.extend(["--ollama-url", kwargs["ollama_url"]])
    if backend == "ollama" and kwargs.get("ollama_auth"):
        cmd.extend(["--ollama-auth", kwargs["ollama_auth"]])
    if backend == "openai" and kwargs.get("openai_key"):
        cmd.extend(["--openai-key", kwargs["openai_key"]])
    if backend == "gemini" and kwargs.get("gemini_key"):
        cmd.extend(["--gemini-key", kwargs["gemini_key"]])
    
    try:
        # Pass current environment to subprocess (includes .env vars)
        env = os.environ.copy()
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
            env=env,
            cwd=str(Path(__file__).parent)
        )
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            
            # Cache the result
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return data
        else:
            print("  ✗ Error: " + result.stderr[:200], file=sys.stderr)
            return None
    except subprocess.TimeoutExpired:
        print("  ✗ Timeout after 180s", file=sys.stderr)
        return None
    except (json.JSONDecodeError, OSError) as e:
        print("  ✗ Exception: " + str(e)[:200], file=sys.stderr)
        return None


def merge_data(original_row: Dict[str, str], extracted_data: Dict) -> Dict[str, str]:
    """Merge extracted data into original row, only filling missing fields."""
    updated_row = original_row.copy()
    changes = []
    
    for extract_key, csv_column in COLUMN_MAPPING.items():
        if csv_column is None:
            continue
        
        # Get extracted value
        extracted_value = extract_nested_value(extracted_data, extract_key)
        if extracted_value is None or is_field_empty(extracted_value):
            continue
        
        # Get original value, initialize new column if doesn't exist
        original_value = updated_row.get(csv_column, "")
        
        # Add new column or fill empty existing column
        if is_field_empty(original_value) or csv_column not in updated_row:
            # Convert lists to comma-separated strings for CSV
            if isinstance(extracted_value, list):
                extracted_value = ", ".join(str(v) for v in extracted_value if v)
            elif isinstance(extracted_value, dict):
                extracted_value = json.dumps(extracted_value)
            
            updated_row[csv_column] = str(extracted_value)
            changes.append(csv_column)
    
    return updated_row, changes


def clean_text_for_excel(text: str) -> str:
    """Clean text to handle encoding issues for Excel."""
    if not isinstance(text, str):
        return text
    
    # First, handle common patterns from corrupted data
    # Pattern: "9:00?AM?�?5:00?PM" -> "9:00 AM - 5:00 PM"
    import re
    text = re.sub(r'(\d{1,2}:\d{2})\?([AP]M)\?[�\ufffd]?\?(\d{1,2}:\d{2})\?([AP]M)', 
                  r'\1 \2 - \3 \4', text)
    
    # Replace common problematic characters
    replacements = {
        '\u2013': '-',  # en dash
        '\u2014': '-',  # em dash
        '\u2018': "'",  # left single quote
        '\u2019': "'",  # right single quote
        '\u201c': '"',  # left double quote
        '\u201d': '"',  # right double quote
        '\u2026': '...',  # ellipsis
        '\xa0': ' ',  # non-breaking space
        '\u00a0': ' ',  # non-breaking space (alternative)
        '\ufffd': '',  # replacement character (�)
        '�': '',  # replacement character (alternative)
        '\x00': '',  # null byte
    }
    
    for old, new in replacements.items():
        text = text.replace(old, new)
    
    # Remove standalone ? that aren't part of real text (but preserve in URLs)
    # Only remove ? if it's followed by letters/spaces (garbage pattern like "9:00?AM")
    # Don't remove ? that's part of URL query strings
    if 'http' not in text.lower():
        # Not a URL, safe to remove all ?
        text = re.sub(r'\?+', ' ', text)
    else:
        # URL detected, only remove ? in garbage patterns, not in query strings
        # Remove patterns like "?AM" or "?PM" but keep "?q=" or "?key="
        text = re.sub(r'\?([AP]M)', r' \1', text)  # Fix time garbage
        text = re.sub(r'(\d)\?([A-Z][a-z])', r'\1 \2', text)  # Fix other garbage
    
    # Clean up multiple spaces
    text = re.sub(r' +', ' ', text)
    
    # Remove any remaining non-printable characters except newlines/tabs
    text = ''.join(char if char.isprintable() or char in '\n\r\t' else ' ' for char in text)
    
    return text.strip()


def export_to_xlsx(data: List[Dict[str, str]], output_path: str, highlight_changes: Dict[int, List[str]]) -> None:
    """Export data to Excel with highlighting for changed cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Resources"
    
    if not data:
        wb.save(output_path)
        return
    
    # Collect all unique column names from all rows (to include new columns)
    all_headers = []
    for row in data:
        for key in row.keys():
            if key not in all_headers:
                all_headers.append(key)
    
    # Write header
    headers = all_headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Write data
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            # Clean text to handle encoding issues
            if isinstance(value, str):
                value = clean_text_for_excel(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Highlight changed cells
            if row_idx - 2 in highlight_changes and header in highlight_changes[row_idx - 2]:
                cell.fill = yellow_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Batch pipeline to enrich service center information"
    )
    parser.add_argument(
        "--input",
        default="data/input/to_be_normalised/enriched_resources.csv",
        help="Input CSV file path"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output XLSX file path (default: auto-generated with timestamp)"
    )
    parser.add_argument(
        "--backend",
        choices=["ollama", "openai", "gemini"],
        default="gemini",
        help="LLM backend to use"
    )
    parser.add_argument(
        "--model",
        default=None,
        help="Model name (optional, uses defaults)"
    )
    parser.add_argument(
        "--enhance-with-websearch",
        action="store_true",
        help="Enable web search enhancement"
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Maximum number of rows to process (for testing)"
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=0,
        help="Row number to start from (0-indexed)"
    )
    parser.add_argument(
        "--cache-dir",
        default=".cache/llm_extractions",
        help="Directory to cache LLM results"
    )
    parser.add_argument(
        "--fields-to-check",
        default="description_short,age_range,organization_type",
        help="Comma-separated list of fields to check for missing data"
    )
    parser.add_argument(
        "--skip-no-website",
        action="store_true",
        help="Skip rows without a website URL"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="Delay between API calls in seconds"
    )
    # Backend-specific arguments
    parser.add_argument("--ollama-url", default=None)
    parser.add_argument("--ollama-auth", default=None)
    parser.add_argument("--openai-key", default=None)
    parser.add_argument("--gemini-key", default=None)
    
    args = parser.parse_args()
    
    # Read input CSV
    print(f"Reading {args.input}...")
    try:
        # Try UTF-8 first, then fall back to other encodings
        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        rows = None
        
        for encoding in encodings_to_try:
            try:
                with open(args.input, "r", encoding=encoding, errors='replace') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                break
            except UnicodeDecodeError:
                continue
        
        if rows is None:
            raise Exception("Could not read CSV with any encoding")
            
    except Exception as e:
        print(f"Error reading CSV: {e}", file=sys.stderr)
        sys.exit(1)
    
    print(f"Loaded {len(rows)} rows")
    
    # Determine output path
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"data/output/enriched_resources_updated_{timestamp}.xlsx"
    
    # Filter rows that need enrichment
    fields_to_check = [f.strip() for f in args.fields_to_check.split(",")]
    rows_to_process = []
    
    for idx, row in enumerate(rows):
        if args.start_row and idx < args.start_row:
            continue
        if args.max_rows and len(rows_to_process) >= args.max_rows:
            break
        
        # Check if row has website
        website = row.get("gmaps_website", "").strip()
        if args.skip_no_website and not website:
            continue
        
        # Check if row needs enrichment
        if needs_enrichment(row, fields_to_check):
            rows_to_process.append((idx, row))
    
    print(f"Found {len(rows_to_process)} rows that need enrichment")
    
    if not rows_to_process:
        print("No rows need enrichment. Exiting.")
        return
    
    # Process rows
    updated_rows = rows.copy()
    highlight_changes = {}
    stats = {
        "processed": 0,
        "enriched": 0,
        "failed": 0,
        "skipped": 0
    }
    
    for process_idx, (original_idx, row) in enumerate(rows_to_process, 1):
        center_name = row.get("gmaps_name", "Unknown")
        website = row.get("gmaps_website", "").strip()
        
        print(f"\n[{process_idx}/{len(rows_to_process)}] Processing: {center_name}")
        print(f"  Website: {website or '(none)'}")
        
        if not website:
            print("  → Skipping (no website)")
            stats["skipped"] += 1
            continue
        
        # Call web_llm_extract
        try:
            extracted_data = call_web_llm_extract(
                center_name=center_name,
                website_url=website,
                backend=args.backend,
                model=args.model,
                enhance_with_websearch=args.enhance_with_websearch,
                cache_dir=args.cache_dir,
                ollama_url=args.ollama_url,
                ollama_auth=args.ollama_auth,
                openai_key=args.openai_key,
                gemini_key=args.gemini_key
            )
            
            if extracted_data and "error" not in extracted_data:
                # Merge data
                updated_row, changes = merge_data(row, extracted_data)
                updated_rows[original_idx] = updated_row
                
                if changes:
                    highlight_changes[original_idx] = changes
                    print("  ✓ Updated fields: " + ", ".join(changes))
                    stats["enriched"] += 1
                else:
                    print("  → No new information to add")
                
                stats["processed"] += 1
            else:
                print("  ✗ Failed to extract data")
                stats["failed"] += 1
        
        except (OSError, ValueError) as e:
            print("  ✗ Exception: " + str(e)[:200])
            stats["failed"] += 1
        
        # Delay between requests
        if process_idx < len(rows_to_process):
            time.sleep(args.delay)
    
    # Export to Excel
    print(f"\n\nExporting to {args.output}...")
    export_to_xlsx(updated_rows, args.output, highlight_changes)
    
    # Print statistics
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total rows in input:     {len(rows)}")
    print(f"Rows needing enrichment: {len(rows_to_process)}")
    print(f"Successfully processed:  {stats['processed']}")
    print(f"Successfully enriched:   {stats['enriched']}")
    print(f"Failed:                  {stats['failed']}")
    print(f"Skipped:                 {stats['skipped']}")
    print(f"\nOutput saved to: {args.output}")
    print("Changed cells are highlighted in yellow")
    print("="*60)


if __name__ == "__main__":
    main()

