#!/usr/bin/env python3
"""
Filter and analyze neurodivergent relevance in enriched data
"""

import argparse
import sys
from pathlib import Path
from collections import Counter

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def analyze_relevance(input_file):
    """Analyze neurodivergent relevance across all resources."""
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Find relevant columns
    headers = [cell.value for cell in ws[1]]
    
    relevance_col = None
    score_col = None
    focus_col = None
    name_col = None
    
    for idx, header in enumerate(headers, 1):
        if header == "is_neurodivergent_related":
            relevance_col = idx
        elif header == "neurodivergent_relevance_score":
            score_col = idx
        elif header == "neurodivergent_focus":
            focus_col = idx
        elif header == "gmaps_name":
            name_col = idx
    
    if not relevance_col:
        print("⚠️  Warning: 'is_neurodivergent_related' column not found.")
        print("   Run the pipeline first to add validation fields.")
        return
    
    # Analyze data
    total = 0
    related = 0
    not_related = 0
    unknown = 0
    
    scores = Counter()
    
    print("\n" + "="*70)
    print("NEURODIVERGENT RELEVANCE ANALYSIS")
    print("="*70)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        
        is_related = str(row[relevance_col - 1]).lower() if relevance_col and row[relevance_col - 1] else ""
        score = str(row[score_col - 1]) if score_col and row[score_col - 1] else "Unknown"
        
        if is_related in ("true", "yes", "1"):
            related += 1
            scores[score] += 1
        elif is_related in ("false", "no", "0"):
            not_related += 1
        else:
            unknown += 1
    
    print(f"\n📊 SUMMARY:")
    print(f"  Total resources:           {total}")
    print(f"  ✅ Neurodivergent-related:  {related} ({related/total*100:.1f}%)")
    print(f"  ❌ Not related:             {not_related} ({not_related/total*100:.1f}%)")
    print(f"  ❓ Unknown/Not validated:   {unknown} ({unknown/total*100:.1f}%)")
    
    if scores:
        print(f"\n📈 RELEVANCE SCORE BREAKDOWN:")
        for score, count in scores.most_common():
            print(f"  {score:15} : {count:4} ({count/related*100:.1f}%)")
    
    return {
        "total": total,
        "related": related,
        "not_related": not_related,
        "unknown": unknown,
        "scores": scores
    }


def filter_by_relevance(input_file, output_file, min_score="Low", include_unknown=False):
    """Filter resources by neurodivergent relevance."""
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Find columns
    headers = [cell.value for cell in ws[1]]
    relevance_col = None
    score_col = None
    
    for idx, header in enumerate(headers, 1):
        if header == "is_neurodivergent_related":
            relevance_col = idx
        elif header == "neurodivergent_relevance_score":
            score_col = idx
    
    if not relevance_col:
        print("❌ Error: Validation columns not found. Run pipeline first.")
        return
    
    # Create new workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    
    # Copy headers
    for col_idx, header in enumerate(headers, 1):
        cell = new_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Score hierarchy
    score_hierarchy = {"High": 3, "Medium": 2, "Low": 1, "None": 0, "Unknown": -1}
    min_score_value = score_hierarchy.get(min_score, 1)
    
    # Filter rows
    filtered_count = 0
    new_row = 2
    
    for row in ws.iter_rows(min_row=2):
        is_related = str(row[relevance_col - 1].value).lower() if row[relevance_col - 1].value else ""
        score = str(row[score_col - 1].value) if score_col and row[score_col - 1].value else "Unknown"
        score_value = score_hierarchy.get(score, -1)
        
        # Include if neurodivergent-related and meets score threshold
        if is_related in ("true", "yes", "1") and score_value >= min_score_value:
            for col_idx, cell in enumerate(row, 1):
                new_ws.cell(row=new_row, column=col_idx, value=cell.value)
            new_row += 1
            filtered_count += 1
        # Include unknown if flag set
        elif include_unknown and (not is_related or is_related == "unknown"):
            for col_idx, cell in enumerate(row, 1):
                new_ws.cell(row=new_row, column=col_idx, value=cell.value)
            new_row += 1
            filtered_count += 1
    
    # Auto-adjust column widths
    for column in new_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        new_ws.column_dimensions[column_letter].width = adjusted_width
    
    new_wb.save(output_file)
    
    print(f"\n✅ Filtered {filtered_count} neurodivergent-related resources")
    print(f"   Saved to: {output_file}")


def show_non_neurodivergent(input_file, limit=20):
    """Show resources that are NOT neurodivergent-related."""
    wb = load_workbook(input_file)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    relevance_col = None
    score_col = None
    name_col = None
    focus_col = None
    
    for idx, header in enumerate(headers, 1):
        if header == "is_neurodivergent_related":
            relevance_col = idx
        elif header == "neurodivergent_relevance_score":
            score_col = idx
        elif header == "gmaps_name":
            name_col = idx
        elif header == "neurodivergent_focus":
            focus_col = idx
    
    print("\n" + "="*70)
    print("NON-NEURODIVERGENT RESOURCES (to review/remove)")
    print("="*70)
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        is_related = str(row[relevance_col - 1]).lower() if relevance_col and row[relevance_col - 1] else ""
        
        if is_related in ("false", "no", "0"):
            count += 1
            if count <= limit:
                name = row[name_col - 1] if name_col else "Unknown"
                score = row[score_col - 1] if score_col else "N/A"
                focus = row[focus_col - 1] if focus_col else "N/A"
                
                print(f"\n{count}. {name}")
                print(f"   Score: {score}")
                print(f"   Focus: {focus}")
    
    if count > limit:
        print(f"\n... and {count - limit} more")
    
    print(f"\nTotal non-neurodivergent: {count}")


def main():
    parser = argparse.ArgumentParser(description="Filter neurodivergent resources")
    parser.add_argument("--input", required=True, help="Input Excel file")
    parser.add_argument("--analyze", action="store_true", help="Analyze relevance distribution")
    parser.add_argument("--filter", action="store_true", help="Filter neurodivergent-related only")
    parser.add_argument("--output", help="Output file for filtered results")
    parser.add_argument("--min-score", choices=["High", "Medium", "Low"], default="Low",
                       help="Minimum relevance score to include")
    parser.add_argument("--show-non-nd", action="store_true", help="Show non-neurodivergent resources")
    parser.add_argument("--include-unknown", action="store_true", 
                       help="Include resources with unknown validation")
    
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        print(f"❌ Error: File not found: {args.input}")
        return
    
    if args.analyze or (not args.filter and not args.show_non_nd):
        analyze_relevance(args.input)
    
    if args.show_non_nd:
        show_non_neurodivergent(args.input)
    
    if args.filter:
        if not args.output:
            args.output = args.input.replace(".xlsx", "_neurodivergent_only.xlsx")
        filter_by_relevance(args.input, args.output, args.min_score, args.include_unknown)
    
    print()


if __name__ == "__main__":
    main()

